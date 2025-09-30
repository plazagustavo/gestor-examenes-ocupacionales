# =============================================================================
# CONVERSOR DE EXCEL - VERSIÓN ORDENADA ALFABÉTICAMENTE
# =============================================================================

import pandas as pd
import os
import sys
import time
import traceback

# Intentar importar openpyxl para formato mejorado
try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, Border, Side
    OPENPYXL_AVAILABLE = True
    print("[OK] openpyxl disponible - formato mejorado activado")
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("[AVISO] openpyxl no disponible - usando formato básico")

# Intentar importar xlwings para autowidth
try:
    import xlwings as xw
    XLWINGS_AVAILABLE = True
    print("[OK] xlwings disponible - autowidth activado")
except ImportError:
    XLWINGS_AVAILABLE = False
    print("[AVISO] xlwings no disponible - sin autowidth")

# Cambiar al directorio donde se encuentra este script
script_dir = os.path.dirname(os.path.abspath(__file__))
os.chdir(script_dir)
print(f"Directorio de trabajo establecido a: {os.getcwd()}")

def apply_autowidth_excel(file_path):
    """
    Aplica autowidth a todas las columnas excepto la columna B y C usando xlwings.
    """
    if not XLWINGS_AVAILABLE:
        print("[AVISO] xlwings no disponible, omitiendo autowidth")
        return True
        
    try:
        print(f"Aplicando autowidth: {os.path.basename(file_path)}")
        
        app = xw.App(visible=False)
        libro = app.books.open(file_path)
        
        for hoja in libro.sheets:
            # Primero establecer anchos fijos para columnas B y C
            hoja.range('B:B').column_width = 38      # Columna B (Empleado)
            hoja.range('C:C').column_width = 11  # Columna C (CUIL)
            
            # Luego autoajuste solo para las otras columnas
            used_range = hoja.used_range
            if used_range:
                last_col = used_range.last_cell.column
                
                for col in range(1, last_col + 1):
                    if col == 2 or col == 3:  # Saltar columnas B y C
                        continue
                    letra = xw.utils.col_name(col)
                    hoja.range(f'{letra}:{letra}').columns.autofit()
        
        # Guardar cambios
        libro.save()
        libro.close()
        app.quit()
        print(f"[OK] Autowidth aplicado correctamente: {os.path.basename(file_path)}")
        return True
        
    except Exception as e:
        print(f"[ERROR] Error aplicando autowidth a {file_path}: {str(e)}")
        try:
            libro.close()
            app.quit()
        except:
            pass
        return False

def extract_company_data_fixed_positions(df):
    """
    Extrae datos de empresa usando posiciones fijas
    """
    try:
        company_data = {
            'Empresa': '',
            'CUIT': '',
            'Contrato': '',
            'Domicilio': '',
            'Localidad': '',
            'Provincia': '',
            'Telefono': '',
            'Email': ''
        }
        
        # Verificar que el DataFrame tenga datos
        if len(df) == 0:
            print("[AVISO] DataFrame vacío")
            return company_data
            
        # Extraer datos con verificación de límites
        if len(df.columns) > 7 and len(df) > 0:
            empresa_val = df.iloc[0, 7]
            company_data['Empresa'] = str(empresa_val) if pd.notna(empresa_val) else ''
            
        if len(df.columns) > 1 and len(df) > 0:
            cuit_val = df.iloc[0, 1]
            company_data['CUIT'] = str(cuit_val) if pd.notna(cuit_val) else ''
            
        if len(df.columns) > 6 and len(df) > 0:
            contrato_val = df.iloc[0, 6]
            company_data['Contrato'] = str(contrato_val) if pd.notna(contrato_val) else ''
            
        if len(df.columns) > 9 and len(df) > 0:
            domicilio_val = df.iloc[0, 9]
            company_data['Domicilio'] = str(domicilio_val) if pd.notna(domicilio_val) else ''
            
        if len(df.columns) > 16 and len(df) > 0:
            localidad_val = df.iloc[0, 16]
            company_data['Localidad'] = str(localidad_val) if pd.notna(localidad_val) else ''
            
        if len(df.columns) > 11 and len(df) > 0:
            provincia_val = df.iloc[0, 11]
            company_data['Provincia'] = str(provincia_val) if pd.notna(provincia_val) else ''
            
        if len(df.columns) > 14 and len(df) > 0:
            telefono_val = df.iloc[0, 14]
            company_data['Telefono'] = str(telefono_val) if pd.notna(telefono_val) else ''
            
        if len(df.columns) > 8 and len(df) > 0:
            email_val = df.iloc[0, 8]
            company_data['Email'] = str(email_val) if pd.notna(email_val) else ''
        
        print(f"Datos de empresa extraídos (posiciones fijas):")
        for key, value in company_data.items():
            print(f"   {key}: {value}")
        
        return company_data
        
    except Exception as e:
        print(f"[ERROR] Error extrayendo datos de empresa: {e}")
        return {
            'Empresa': '',
            'CUIT': '',
            'Contrato': '',
            'Domicilio': '',
            'Localidad': '',
            'Provincia': '',
            'Telefono': '',
            'Email': ''
        }

def process_all_patients(df):
    """
    Procesa todos los pacientes en el Excel y asigna números únicos
    MISMA LÓGICA que fix_de_id.py - MÉTODO PRINCIPAL DE EXTRACCIÓN
    """
    try:
        # COLUMNAS:
        # CUIL: C (índice 2)
        # Apellido y Nombre: P (índice 15)
        # Descripción: E (índice 4)
        
        cuil_col = 2  # Columna C
        apellido_nombre_col = 15  # Columna P
        descripcion_col = 4  # Columna E
        
        # Diccionarios para almacenar información
        patient_numbers = {}
        patient_info = {}
        
        # Lista para almacenar los CUILs únicos
        unique_cuils = []
        
        print(f"Procesando {len(df)} filas...")
        
        # Recorrer todas las filas excepto la primera (encabezado)
        for i in range(1, len(df)):
            try:
                # Verificamos que los índices estén dentro de los límites
                if (cuil_col < len(df.columns) and 
                    apellido_nombre_col < len(df.columns) and 
                    descripcion_col < len(df.columns)):
                    
                    # Obtener CUIL
                    cuil_value = df.iloc[i, cuil_col]
                    cuil = str(cuil_value).strip() if pd.notna(cuil_value) else ""
                    
                    # Obtener nombre
                    nombre_value = df.iloc[i, apellido_nombre_col]
                    nombre = str(nombre_value).strip() if pd.notna(nombre_value) else ""
                    
                    # Obtener descripción
                    desc_value = df.iloc[i, descripcion_col]
                    descripcion = str(desc_value).strip() if pd.notna(desc_value) else ""
                    
                    # Si tenemos un CUIL válido y no es un encabezado
                    if cuil and cuil.lower() != "cuil" and len(cuil) > 5:
                        # Si es la primera vez que vemos este CUIL
                        if cuil not in unique_cuils:
                            unique_cuils.append(cuil)
                            
                            # Guardar información del paciente
                            patient_info[cuil] = {
                                "nombre": nombre,
                                "estudios": [descripcion] if descripcion else []
                            }
                        else:
                            # Agregar este estudio a la lista de estudios del paciente
                            if descripcion and descripcion not in patient_info[cuil]["estudios"]:
                                patient_info[cuil]["estudios"].append(descripcion)
            except Exception as e:
                print(f"[AVISO] Error procesando fila {i}: {e}")
                continue
        
        # Ordenar pacientes alfabéticamente por nombre antes de asignar números
        cuil_nombre_pairs = [(cuil, info["nombre"]) for cuil, info in patient_info.items()]
        cuil_nombre_pairs.sort(key=lambda x: x[1].upper())
        
        # Asignar números a los CUILs ordenados alfabéticamente, empezando desde 1
        for i, (cuil, nombre) in enumerate(cuil_nombre_pairs):
            patient_numbers[cuil] = i + 1
        
        print(f"Información de procesamiento:")
        print(f"   - Total de filas procesadas: {len(df) - 1}")
        print(f"   - Pacientes únicos encontrados: {len(patient_numbers)}")
        print(f"   - Ordenados alfabéticamente por nombre")
        
        # Mostrar algunos ejemplos para verificación
        print(f"Primeros 5 pacientes (ordenados A-Z):")
        for i, (cuil, nombre) in enumerate(cuil_nombre_pairs[:5]):
            num = patient_numbers[cuil]
            info = patient_info[cuil]
            estudios = ", ".join(info["estudios"]) if info["estudios"] else "Sin estudios"
            print(f"   {num}. {nombre} (CUIL: {cuil}) - Estudios: {estudios}")
        
        return patient_info, patient_numbers
        
    except Exception as e:
        print(f"[ERROR] Error procesando pacientes: {e}")
        return {}, {}

def process_excel_file_with_openpyxl(input_file, output_file, company_data, final_employees_data, exams_list, exam_count, patient_numbers):
    """
    Crea el archivo Excel con formato profesional usando openpyxl
    ORDENANDO ALFABÉTICAMENTE POR NOMBRE
    """
    try:
        # Create a new workbook
        wb = Workbook()
        ws = wb.active
        
        # Write company data
        company_fields = [
            ('Proceso', ''),
            ('Empresa', company_data['Empresa']),
            ('CUIT', company_data['CUIT']),
            ('Contrato', company_data['Contrato']),
            ('Domicilio', company_data['Domicilio']),
            ('Localidad', company_data['Localidad']),
            ('Provincia', company_data['Provincia']),
            ('Telefono', company_data['Telefono']),
            ('Contacto', ''),
            ('Email', company_data['Email'])
        ]
        
        for i, (field, value) in enumerate(company_fields, start=2):
            ws.cell(row=i, column=2, value=field).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=i, column=3, value=value).alignment = Alignment(horizontal='left', vertical='center')
        
        # Leave a blank row
        current_row = len(company_fields) + 3
        
        # Define border style (thin black border)
        thin_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        # Write headers
        headers = ['Id', 'Empleado', 'CUIL'] + exams_list
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=current_row, column=col_idx, value=header)
            if col_idx > 3:  # For exams
                cell.alignment = Alignment(text_rotation=90, vertical='bottom', horizontal='center', wrap_text=True)
            else:
                cell.alignment = Alignment(vertical='bottom', horizontal='center', wrap_text=True)
            cell.font = Font(bold=True)
            cell.border = thin_border
        
        # Adjust column widths
        column_widths = [4.86, 53, 13.29] + [10.57] * len(exams_list)
        for idx, width in enumerate(column_widths, start=1):
            col_letter = ws.cell(row=1, column=idx).column_letter
            ws.column_dimensions[col_letter].width = width
        
        # Adjust header row height
        ws.row_dimensions[current_row].height = 120
        
        # Write employee data - ordenado alfabéticamente por nombre
        row_idx = current_row + 1
        last_employee_row = row_idx
        
        sorted_employees = sorted(final_employees_data.items(), key=lambda x: x[1]['name'].upper())
        
        for idx, (cuil, data) in enumerate(sorted_employees, start=1):
            patient_number = patient_numbers[cuil]
            
            ws.cell(row=row_idx, column=1, value=patient_number).border = thin_border
            ws.cell(row=row_idx, column=2, value=data['name']).border = thin_border
            ws.cell(row=row_idx, column=3, value=data['cuil']).border = thin_border
            
            # Mark exams with X
            for col_idx, exam in enumerate(exams_list, start=4):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = thin_border
                
                if any(e.strip() == exam.strip() for e in data['exams']):
                    cell.value = "X"
                    cell.alignment = Alignment(horizontal='center', vertical='center')
            
            row_idx += 1
            last_employee_row = row_idx - 1
        
        # Add blank row
        row_idx += 1
        
        # Add exam count
        for exam, count in exam_count.items():
            ws.cell(row=row_idx, column=1, value=count)
            ws.cell(row=row_idx, column=2, value=exam)
            row_idx += 1
        
        # Save the file
        wb.save(output_file)
        return True, last_employee_row
        
    except Exception as e:
        print(f"[ERROR] Error creando archivo con openpyxl: {e}")
        return False, None

def process_excel_file_with_pandas(input_file, output_file, company_data, final_employees_data, exams_list, exam_count, patient_numbers):
    """
    Crea el archivo Excel con formato básico usando solo pandas (fallback)
    ORDENANDO ALFABÉTICAMENTE POR NOMBRE
    """
    try:
        # Crear un nuevo DataFrame para el resultado
        company_rows = []
        company_fields = [
            ('Proceso', ''),
            ('Empresa', company_data['Empresa']),
            ('CUIT', company_data['CUIT']),
            ('Contrato', company_data['Contrato']),
            ('Domicilio', company_data['Domicilio']),
            ('Localidad', company_data['Localidad']),
            ('Provincia', company_data['Provincia']),
            ('Telefono', company_data['Telefono']),
            ('Contacto', ''),
            ('Email', company_data['Email'])
        ]
        
        for field, value in company_fields:
            row = [''] * (3 + len(exams_list))
            row[1] = field
            row[2] = value
            company_rows.append(row)
        
        # Agregar fila en blanco
        company_rows.append([''] * (3 + len(exams_list)))
        
        # Crear encabezados
        headers = ['Id', 'Empleado', 'CUIL'] + exams_list
        
        # Ordenar empleados alfabéticamente por nombre
        sorted_employees = sorted(final_employees_data.items(), key=lambda x: x[1]['name'].upper())
        
        # Crear filas de empleados
        employee_rows = []
        for idx, (cuil, data) in enumerate(sorted_employees, start=1):
            row = [''] * (3 + len(exams_list))
            patient_number = patient_numbers[cuil]
            row[0] = patient_number
            row[1] = data['name']
            row[2] = data['cuil']
            
            # Marcar exámenes con X
            for i, exam in enumerate(exams_list):
                if any(e.strip() == exam.strip() for e in data['exams']):
                    row[3 + i] = 'X'
            
            employee_rows.append(row)
        
        # Agregar fila en blanco
        employee_rows.append([''] * (3 + len(exams_list)))
        
        # Agregar recuento de exámenes
        exam_rows = []
        for exam, count in exam_count.items():
            row = [''] * (3 + len(exams_list))
            row[0] = count
            row[1] = exam
            exam_rows.append(row)
        
        # Combinar todas las filas
        all_rows = company_rows + [headers] + employee_rows + exam_rows
        
        # Crear DataFrame final
        result_df = pd.DataFrame(all_rows)
        
        # Guardar como Excel
        result_df.to_excel(output_file, index=False, header=False)
        return True, len(employee_rows)
        
    except Exception as e:
        print(f"[ERROR] Error creando archivo con pandas: {e}")
        return False, None

def process_excel_file(input_file, output_file):
    """
    Process the Excel file and generate a new formatted Excel file.
    Ordena alfabéticamente por nombre
    """
    print(f"Procesando archivo: {input_file}")
    
    try:
        # Verificar que el archivo existe
        if not os.path.exists(input_file):
            print(f"[ERROR] El archivo {input_file} no existe")
            return False
        
        # Read the Excel file with header for company data
        print("Leyendo archivo Excel...")
        df_with_header = pd.read_excel(input_file)
        
        # Read the Excel file without header for patient processing
        df_no_header = pd.read_excel(input_file, header=None)
        
        print(f"Archivo leído: {len(df_with_header)} filas, {len(df_with_header.columns)} columnas")
        
        # Extracción de datos de empresa
        company_data = extract_company_data_fixed_positions(df_with_header)
        
        # Procesamiento de pacientes con ordenamiento alfabético
        print("\nUsando lógica de fix_de_id.py con ordenamiento alfabético...")
        patient_info, patient_numbers = process_all_patients(df_no_header)
        
        # Convertir la información de pacientes al formato esperado
        final_employees_data = {}
        exams_set = set()
        exam_count = {}
        
        for cuil, info in patient_info.items():
            final_employees_data[cuil] = {
                'name': info['nombre'],
                'cuil': cuil,
                'exams': info['estudios']
            }
            
            # Contar exámenes
            for exam in info['estudios']:
                if exam:
                    exams_set.add(exam.strip())
                    if exam.strip() not in exam_count:
                        exam_count[exam.strip()] = 0
                    exam_count[exam.strip()] += 1
        
        print(f"\n[OK] Total de empleados extraídos: {len(final_employees_data)}")
        print(f"[OK] Método de extracción: process_all_patients con ordenamiento alfabético")
        
        # Define the preferred order for exams
        preferred_order = [
            "EXAMEN CLINICO",
            "AUDIOMETRIA",
            "ESPIROMETRIA",
            "CUESTIONARIO OSTEOARTICULAR COLUMNA LUMBOSACRA",
            "CUESTIONARIO DE SEGMENTOS COMPROMETIDOS",
            "RX",
            "RX DE TORAX"
        ]
        
        # Sort exams according to preferred order
        exams_list = []
        
        # First add the preferred exams in the specified order (if they exist in the data)
        for exam in preferred_order:
            matching_exams = [e for e in exams_set if exam.upper() in e.upper()]
            for matching_exam in matching_exams:
                if matching_exam in exams_set:
                    exams_list.append(matching_exam)
                    exams_set.remove(matching_exam)
        
        # Then add all remaining exams alphabetically
        remaining_exams = sorted(list(exams_set))
        exams_list.extend(remaining_exams)
        
        # Crear archivo con el mejor formato disponible
        if OPENPYXL_AVAILABLE:
            print("Usando openpyxl para formato profesional...")
            result = process_excel_file_with_openpyxl(input_file, output_file, company_data, final_employees_data, exams_list, exam_count, patient_numbers)
        else:
            print("Usando pandas para formato básico...")
            result = process_excel_file_with_pandas(input_file, output_file, company_data, final_employees_data, exams_list, exam_count, patient_numbers)
        
        if result[0]:
            print(f"[OK] Archivo guardado como: {output_file}")
            
            # Show summary
            print(f"\nRESUMEN FINAL:")
            print(f"   Empresa: {company_data['Empresa']}")
            print(f"   Cantidad de empleados: {len(final_employees_data)}")
            print(f"   Cantidad de exámenes: {len(exams_list)}")
            print(f"   Pacientes ordenados alfabéticamente (A-Z)")
            print(f"   Cada paciente mantiene sus estudios correspondientes")
            
            return True, result[1]
        else:
            return False
    
    except Exception as e:
        print(f"[ERROR] Error al procesar el archivo: {str(e)}")
        traceback.print_exc()
        return False

def main():
    print("=" * 60)
    print("CONVERSOR DE EXCEL AUTOMÁTICO - VERSIÓN ORDENADA A-Z")
    print("=" * 60)
    print("Procesamiento automático activado")
    print("NUEVO: Pacientes ordenados alfabéticamente A-Z")
    print("Cada paciente mantiene sus estudios correspondientes")
    print("Extracción de datos de empresa con método fijo")
    print("Verificación de pacientes por CUIL")
    print("Formato profesional con openpyxl (si está disponible)")
    print("Autowidth automático con xlwings (si está disponible)")
    print("=" * 60)
    
    # Mostrar el directorio actual
    carpeta = os.path.dirname(os.path.abspath(__file__))
    print(f"Directorio actual: {carpeta}")
    
    # Buscar todos los archivos xlsx en la carpeta
    archivos_xlsx = []
    try:
        for archivo in os.listdir(carpeta):
            if archivo.endswith('.xlsx') and not archivo.startswith('~$') and not archivo.startswith('output_'):
                archivos_xlsx.append(archivo)
        
        # También buscar archivos .xls
        for archivo in os.listdir(carpeta):
            if archivo.endswith('.xls') and not archivo.startswith('~$') and not archivo.startswith('output_'):
                archivos_xlsx.append(archivo)
    except Exception as e:
        print(f"[ERROR] Error listando archivos: {e}")
        input("\nPresiona Enter para cerrar el programa...")
        return
    
    if not archivos_xlsx:
        print("[ERROR] No se encontraron archivos .xlsx o .xls para procesar en la carpeta.")
        input("\nPresiona Enter para cerrar el programa...")
        return
    
    print(f"Archivos encontrados: {len(archivos_xlsx)}")
    for archivo in archivos_xlsx:
        print(f"   > {archivo}")
    
    print("\n" + "=" * 60)
    print("INICIANDO PROCESAMIENTO AUTOMÁTICO")
    print("=" * 60)
    
    archivos_procesados = 0
    archivos_con_error = 0
    start_time_total = time.time()
    
    # Procesar cada archivo automáticamente
    for archivo in archivos_xlsx:
        print(f"\nProcesando: {archivo}")
        print("-" * 40)
        
        start_time = time.time()
        input_file = os.path.join(carpeta, archivo)
        
        # Generar nombre del archivo de salida
        nombre_sin_extension = os.path.splitext(archivo)[0]
        output_file = os.path.join(carpeta, f"output_sorted_{nombre_sin_extension}.xlsx")
        
        # Procesar el archivo
        result = process_excel_file(input_file, output_file)
        if result and result[0]:
            success, last_employee_row = result
            # Aplicar autowidth al archivo generado
            if apply_autowidth_excel(output_file):
                elapsed_time = time.time() - start_time
                print(f"Archivo procesado exitosamente!")
                print(f"Tiempo: {elapsed_time:.2f} segundos")
                archivos_procesados += 1
            else:
                print(f"[AVISO] Archivo procesado pero falló el autowidth")
                archivos_procesados += 1
        else:
            print(f"[ERROR] Error al procesar el archivo")
            archivos_con_error += 1
    
    # Resumen final
    elapsed_time_total = time.time() - start_time_total
    print("\n" + "=" * 60)
    print("RESUMEN FINAL")
    print("=" * 60)
    print(f"Total de archivos encontrados: {len(archivos_xlsx)}")
    print(f"Archivos procesados exitosamente: {archivos_procesados}")
    print(f"Archivos con errores: {archivos_con_error}")
    print(f"Tiempo total de procesamiento: {elapsed_time_total:.2f} segundos")
    print("=" * 60)
    
    if archivos_procesados > 0:
        print(f"Procesamiento automático completado!")
        print(f"Los archivos de salida están en la misma carpeta con prefijo 'output_'")
        print(f"Formato profesional aplicado")
        print(f"PACIENTES ORDENADOS ALFABÉTICAMENTE A-Z")
        print(f"Cada paciente mantiene sus estudios correspondientes")
    
    print("=" * 60)
    
    input("\nPresiona Enter para cerrar el programa...")

if __name__ == "__main__":
    main()